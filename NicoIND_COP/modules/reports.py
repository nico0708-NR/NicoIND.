"""
NicoIND — modules/reports.py
Generación de reportes PDF y Excel empresariales.
Requiere: reportlab, openpyxl
"""
from flask import Blueprint, render_template, Response
from flask_login import login_required, current_user
from datetime import datetime
from models import db, Product, Employee, FinancialTransaction, SalesOrder
import io

reports_bp = Blueprint("reports", __name__)

def cid(): return current_user.company_id

# ── Imports opcionales ───────────────────────────────────────
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, cm
    REPORTLAB = True
except ImportError:
    REPORTLAB = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL = True
except ImportError:
    OPENPYXL = False


def _pdf_header_style():
    s = getSampleStyleSheet()
    title_style = ParagraphStyle("NITitle", parent=s["Heading1"],
        textColor=colors.HexColor("#7c3aed"), fontSize=16,
        spaceAfter=6, fontName="Helvetica-Bold")
    sub_style = ParagraphStyle("NISub", parent=s["Normal"],
        textColor=colors.HexColor("#6b7280"), fontSize=9, spaceAfter=16)
    return s, title_style, sub_style


def _pdf_table_style(header_color="#7c3aed"):
    return TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  colors.HexColor(header_color)),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1), 8),
        ("ALIGN",         (0, 0), (-1, 0),  "CENTER"),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.HexColor("#f5f3ff"), colors.white]),
        ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d5db")),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING",   (0, 0), (-1, -1), 6),
    ])


# ════════════════════════════════════════════════════════════
# RUTAS
# ════════════════════════════════════════════════════════════

@reports_bp.route("/reports")
@login_required
def index():
    return render_template("reports/index.html",
                           REPORTLAB=REPORTLAB, OPENPYXL=OPENPYXL)


# ── INVENTARIO PDF ───────────────────────────────────────────
@reports_bp.route("/reports/inventory/pdf")
@login_required
def inventory_pdf():
    if not REPORTLAB:
        return "ReportLab no instalado. Ejecuta: pip install reportlab", 500

    products = Product.query.filter_by(company_id=cid(), active=True)\
                             .order_by(Product.category, Product.name).all()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
          rightMargin=1.5*cm, leftMargin=1.5*cm,
          topMargin=1.5*cm, bottomMargin=1.5*cm)
    s, ts, ss = _pdf_header_style()
    elems = [
        Paragraph("NicoIND — Reporte de Inventario", ts),
        Paragraph(f"Empresa: {current_user.company.name}  |  "
                  f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ss),
        HRFlowable(width="100%", thickness=1, color=colors.HexColor("#7c3aed")),
        Spacer(1, 10),
    ]
    headers = ["Código","Producto","Categoría","Unidad","Stock","Mín","Costo","Precio","Valor"]
    data    = [headers]
    for p in products:
        data.append([p.code or "-", p.name[:28], p.category, p.unit,
                     f"{p.stock:.1f}", f"{p.min_stock:.0f}",
                     f"${p.cost:.2f}", f"${p.price:.2f}", f"${p.stock_value:.2f}"])
    t = Table(data, repeatRows=1,
              colWidths=[1.4*cm,6*cm,3*cm,1.6*cm,1.5*cm,1.2*cm,1.8*cm,1.8*cm,2*cm])
    t.setStyle(_pdf_table_style())
    elems.append(t)
    total_val = sum(p.stock_value for p in products)
    low_count = sum(1 for p in products if p.stock_status in ("low","out"))
    elems += [
        Spacer(1, 14),
        Paragraph(f"<b>Total productos: {len(products)}  |  "
                  f"Stock bajo/agotado: {low_count}  |  "
                  f"Valor total: ${total_val:,.2f}</b>", s["Normal"]),
    ]
    doc.build(elems)
    buf.seek(0)
    return Response(buf.getvalue(), mimetype="application/pdf",
        headers={"Content-Disposition": "attachment;filename=inventario_nicond.pdf"})


# ── INVENTARIO EXCEL ─────────────────────────────────────────
@reports_bp.route("/reports/inventory/excel")
@login_required
def inventory_excel():
    if not OPENPYXL:
        return "OpenPyXL no instalado. Ejecuta: pip install openpyxl", 500

    products = Product.query.filter_by(company_id=cid(), active=True)\
                             .order_by(Product.category, Product.name).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    hfill = PatternFill("solid", fgColor="7C3AED")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"))

    headers = ["Código","Nombre","Categoría","Unidad","Stock Actual",
               "Stock Mín","Costo","Precio","Valor Stock","Estado"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill, cell.font = hfill, hfont
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    smap = {"ok":"✅ OK","low":"⚠️ Bajo","out":"🔴 Agotado","over":"📦 Exceso"}
    for r, p in enumerate(products, 2):
        row_data = [p.code or "-", p.name, p.category, p.unit,
                    p.stock, p.min_stock, p.cost, p.price,
                    p.stock_value, smap.get(p.stock_status, "OK")]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F5F3FF")

    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 10
    for col in ["E","F","G","H","I"]: ws.column_dimensions[col].width = 14

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return Response(buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=inventario_nicond.xlsx"})


# ── FINANZAS PDF ─────────────────────────────────────────────
@reports_bp.route("/reports/finance/pdf")
@login_required
def finance_pdf():
    if not REPORTLAB:
        return "ReportLab no instalado", 500

    txs = FinancialTransaction.query.filter_by(company_id=cid())\
          .order_by(FinancialTransaction.date.desc()).limit(100).all()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
          rightMargin=1.5*cm, leftMargin=1.5*cm,
          topMargin=1.5*cm, bottomMargin=1.5*cm)
    s, ts, ss = _pdf_header_style()
    total_inc = sum(t.amount for t in txs if t.transaction_type == "income")
    total_exp = sum(t.amount for t in txs if t.transaction_type == "expense")
    profit    = total_inc - total_exp

    summary = Table([
        ["Concepto","Monto"],
        ["Total Ingresos",  f"${total_inc:,.2f}"],
        ["Total Gastos",    f"${total_exp:,.2f}"],
        ["Utilidad Neta",   f"${profit:,.2f}"],
    ], colWidths=[6*cm, 3*cm])
    summary.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#0ea5e9")),
        ("TEXTCOLOR", (0,0),(-1,0),colors.white),
        ("FONTNAME",  (0,0),(-1,-1),"Helvetica-Bold"),
        ("FONTSIZE",  (0,0),(-1,-1),9),
        ("GRID",      (0,0),(-1,-1),0.4,colors.HexColor("#d1d5db")),
        ("BACKGROUND",(0,3),(-1,3),colors.HexColor("#d1fae5")),
        ("TEXTCOLOR", (0,3),(-1,3),colors.HexColor("#065f46")),
    ]))
    detail_data = [["Fecha","Tipo","Categoría","Descripción","Monto"]]
    for t in txs[:80]:
        detail_data.append([
            t.date.strftime("%d/%m/%Y"),
            "Ingreso" if t.transaction_type=="income" else "Gasto",
            t.category or "-",
            (t.description or "")[:32],
            f"${t.amount:,.2f}"
        ])
    detail_tbl = Table(detail_data, repeatRows=1,
                       colWidths=[2.2*cm,2*cm,3.5*cm,6.5*cm,2.2*cm])
    detail_tbl.setStyle(_pdf_table_style("#0ea5e9"))
    doc.build([
        Paragraph("NicoIND — Reporte Financiero", ts),
        Paragraph(f"Empresa: {current_user.company.name}  |  "
                  f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ss),
        Spacer(1,8), summary, Spacer(1,16), detail_tbl,
    ])
    buf.seek(0)
    return Response(buf.getvalue(), mimetype="application/pdf",
        headers={"Content-Disposition": "attachment;filename=finanzas_nicond.pdf"})


# ── FINANZAS EXCEL ───────────────────────────────────────────
@reports_bp.route("/reports/finance/excel")
@login_required
def finance_excel():
    if not OPENPYXL:
        return "OpenPyXL no instalado", 500

    txs = FinancialTransaction.query.filter_by(company_id=cid())\
          .order_by(FinancialTransaction.date.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Transacciones"

    hfill = PatternFill("solid", fgColor="0EA5E9")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    for c, h in enumerate(["Fecha","Tipo","Categoría","Descripción","Monto","Referencia"],1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill, cell.font = hfill, hfont
        cell.alignment = Alignment(horizontal="center")

    for r, t in enumerate(txs, 2):
        ws.append([t.date.strftime("%d/%m/%Y"),
                   "Ingreso" if t.transaction_type=="income" else "Gasto",
                   t.category or "-", t.description or "-", t.amount, t.reference or "-"])
        if r % 2 == 0:
            for c in range(1, 7):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="E0F2FE")

    ws2 = wb.create_sheet("Resumen")
    total_inc = sum(t.amount for t in txs if t.transaction_type == "income")
    total_exp = sum(t.amount for t in txs if t.transaction_type == "expense")
    ws2.append(["Concepto","Monto"])
    ws2.append(["Total Ingresos", total_inc])
    ws2.append(["Total Gastos",   total_exp])
    ws2.append(["Utilidad Neta",  total_inc - total_exp])
    ws2.append(["Margen %", f"{(total_inc-total_exp)/total_inc*100:.1f}%" if total_inc else "0%"])

    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["C"].width = 20
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return Response(buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=finanzas_nicond.xlsx"})


# ── RRHH EXCEL ───────────────────────────────────────────────
@reports_bp.route("/reports/hr/excel")
@login_required
def hr_excel():
    if not OPENPYXL:
        return "OpenPyXL no instalado", 500

    employees = Employee.query.filter_by(company_id=cid())\
                              .order_by(Employee.full_name).all()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Empleados"
    hfill = PatternFill("solid", fgColor="10B981")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    for c, h in enumerate(["Código","Nombre","Departamento","Cargo",
                            "Salario","F. Ingreso","Estado","Email","Teléfono"],1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill, cell.font = hfill, hfont

    for r, e in enumerate(employees, 2):
        dept = e.department.name if e.department else "—"
        ws.append([e.code or "-", e.full_name, dept, e.position,
                   e.salary, e.hire_date.isoformat() if e.hire_date else "",
                   e.status, e.email or "", e.phone or ""])
        if r % 2 == 0:
            for c in range(1, 10):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="D1FAE5")

    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 25
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return Response(buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=empleados_nicond.xlsx"})


# ── VENTAS EXCEL ─────────────────────────────────────────────
@reports_bp.route("/reports/sales/excel")
@login_required
def sales_excel():
    if not OPENPYXL:
        return "OpenPyXL no instalado", 500

    orders = SalesOrder.query.filter_by(company_id=cid())\
             .order_by(SalesOrder.created_at.desc()).limit(500).all()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Ventas"
    hfill = PatternFill("solid", fgColor="F59E0B")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    for c, h in enumerate(["#Orden","Cliente","Total","Estado","Fecha"],1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill, cell.font = hfill, hfont

    total = 0
    for r, o in enumerate(orders, 2):
        ws.append([o.order_number, o.customer, o.total, o.status,
                   o.created_at.strftime("%d/%m/%Y %H:%M")])
        total += o.total
        if r % 2 == 0:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="FEF3C7")

    ws.cell(row=len(orders)+3, column=1, value="TOTAL")
    ws.cell(row=len(orders)+3, column=3, value=round(total, 2)).font = Font(bold=True)
    ws.column_dimensions["B"].width = 30
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return Response(buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=ventas_nicond.xlsx"})
